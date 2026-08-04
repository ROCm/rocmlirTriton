# Part of the MLIR Project, under the Apache License v2.0 with LLVM Exceptions.
# See https://llvm.org/LICENSE.txt for license information.
# SPDX-License-Identifier: Apache-2.0 WITH LLVM-exception
"""Generate a color-coded, per-architecture attention comparison workbook."""

import argparse
import csv
import math
import os
import re
import sys
import tempfile
from pathlib import Path
from typing import Dict, List, Mapping, Optional, Sequence, Set

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from attentionPerfUtils import canonical_config
from compareAttentionPerformance import OUTPUT_FIELDS

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
MAGENTA_FILL = PatternFill("solid", fgColor="FF00FF")
INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
NUMERIC_COLUMNS = {"Base Samples", "Candidate Samples", "Base TFlops", "Candidate TFlops", "% Diff"}


def read_comparisons(paths: Sequence[Path]) -> List[Dict[str, str]]:
    rows = []
    for path in paths:
        with path.open("r", encoding="utf-8", newline="") as stream:
            reader = csv.DictReader(stream)
            missing = set(OUTPUT_FIELDS) - set(reader.fieldnames or [])
            if missing:
                raise ValueError(f"{path} is missing columns: {sorted(missing)}")
            rows.extend(reader)
    return rows


def sanitize_sheet_name(name: str, used: Set[str]) -> str:
    base = INVALID_SHEET_CHARS.sub("_", name).strip("'") or "unknown"
    base = base[:31]
    candidate = base
    suffix = 1
    while candidate in used:
        suffix_text = f"_{suffix}"
        candidate = f"{base[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used.add(candidate)
    return candidate


def _number(value: str) -> Optional[float]:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _converted_row(row: Mapping[str, str]) -> List:
    result = []
    for field in OUTPUT_FIELDS:
        value = row[field]
        if field in NUMERIC_COLUMNS:
            number = _number(value)
            result.append(number if number is not None else value)
        else:
            result.append(value)
    return result


def _comparison_fill(diff: Optional[float], threshold: float):
    if diff is None:
        return MAGENTA_FILL, MAGENTA_FILL, MAGENTA_FILL
    if diff > threshold:
        return RED_FILL, GREEN_FILL, GREEN_FILL
    if diff < -threshold:
        return GREEN_FILL, RED_FILL, RED_FILL
    return YELLOW_FILL, YELLOW_FILL, YELLOW_FILL


def create_workbook(rows: Sequence[Mapping[str, str]], tie_threshold: float = 0.5) -> Workbook:
    if tie_threshold < 0:
        raise ValueError("tie threshold must be nonnegative")
    if not rows:
        raise ValueError("No comparison rows were provided")

    grouped: Dict[str, List[Mapping[str, str]]] = {}
    seen = set()
    revision_pairs = {(row["Base SHA"], row["Candidate SHA"]) for row in rows}
    if len(revision_pairs) != 1:
        raise ValueError("Workbook inputs must contain exactly one base/candidate revision pair")
    for row in rows:
        key = (row["Chip"], canonical_config(row["Config"]))
        if key in seen:
            raise ValueError(f"Duplicate comparison row: {key}")
        seen.add(key)
        grouped.setdefault(row["Chip"], []).append(row)
    experiment_shapes = {
        arch: {
            canonical_config(row["Config"]):
                (row["RocmlirGenFlags"], row["Base Samples"], row["Candidate Samples"])
            for row in arch_rows
        } for arch, arch_rows in grouped.items()
    }
    for shape in experiment_shapes.values():
        for _, base_samples, candidate_samples in shape.values():
            try:
                base_count = int(base_samples)
                candidate_count = int(candidate_samples)
            except ValueError as error:
                raise ValueError("Sample counts must be positive integers") from error
            if base_count < 1 or base_count != candidate_count:
                raise ValueError("Base and candidate sample counts must match and be positive")
    expected_shape = next(iter(experiment_shapes.values()))
    if any(shape != expected_shape for shape in experiment_shapes.values()):
        raise ValueError(
            "Every architecture must contain the same configs, runtime flags, and sample counts")

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "rocmlirTriton Attention Performance Comparison"
    workbook.properties.subject = "Base versus candidate quick-tuned attention performance"
    used_names: Set[str] = set()

    for arch in sorted(grouped):
        sheet = workbook.create_sheet(sanitize_sheet_name(arch, used_names))
        sheet.append(OUTPUT_FIELDS)
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        base_column = OUTPUT_FIELDS.index("Base TFlops") + 1
        candidate_column = OUTPUT_FIELDS.index("Candidate TFlops") + 1
        diff_column = OUTPUT_FIELDS.index("% Diff") + 1
        for row in sorted(grouped[arch], key=lambda value: value["Config"]):
            sheet.append(_converted_row(row))
            excel_row = sheet.max_row
            diff = _number(row["% Diff"])
            fills = _comparison_fill(diff, tie_threshold)
            sheet.cell(excel_row, base_column).fill = fills[0]
            sheet.cell(excel_row, candidate_column).fill = fills[1]
            sheet.cell(excel_row, diff_column).fill = fills[2]
            sheet.cell(excel_row, base_column).number_format = "0.000"
            sheet.cell(excel_row, candidate_column).number_format = "0.000"
            sheet.cell(excel_row, diff_column).number_format = '0.00"%"'

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_index, field in enumerate(OUTPUT_FIELDS, 1):
            values = [
                str(sheet.cell(row, column_index).value or "")
                for row in range(1, sheet.max_row + 1)
            ]
            width = min(max(len(value) for value in values) + 2, 80 if field == "Config" else 40)
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        sheet.row_dimensions[1].height = 24
    return workbook


def save_workbook(workbook: Workbook, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(prefix=f".{output.name}.",
                                             suffix=".xlsx",
                                             dir=str(output.parent))
    os.close(descriptor)
    try:
        workbook.save(temporary)
        os.replace(temporary, output)
    except BaseException:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass
        raise


def parse_arguments(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, nargs="+", required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--tie-threshold", type=float, default=0.5)
    return parser.parse_args(argv)


def main(argv=None) -> int:
    args = parse_arguments(argv)
    rows = read_comparisons(args.input)
    workbook = create_workbook(rows, args.tie_threshold)
    save_workbook(workbook, args.output)
    print(f"Wrote {len(workbook.sheetnames)} architecture sheets to {args.output}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())

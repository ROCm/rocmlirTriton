# Part of the MLIR Project, under the Apache License v2.0 with LLVM Exceptions.
# See https://llvm.org/LICENSE.txt for license information.
# SPDX-License-Identifier: Apache-2.0 WITH LLVM-exception
#
# RUN: env PYTHONPATH=%S/../../../utils/performance %python %s

import csv
import math
import os
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

import compareAttentionPerformance
import generateAttentionPerformanceWorkbook

comparison = compareAttentionPerformance
spreadsheet = generateAttentionPerformanceWorkbook


def comparison_row(chip="gfx90a", config="-g 1", diff=10.0):
    row = {field: "" for field in comparison.OUTPUT_FIELDS}
    row.update({
        "Chip": chip,
        "Config": config,
        "Base SHA": "base",
        "Candidate SHA": "candidate",
        "Base PerfConfig": "base-p",
        "Candidate PerfConfig": "candidate-p",
        "Base Samples": "3",
        "Candidate Samples": "3",
        "Base TFlops": "10",
        "Candidate TFlops": str(10 * (1 + diff / 100)),
        "% Diff": str(diff),
    })
    return row


class GenerateAttentionWorkbookTest(unittest.TestCase):

    def test_read_and_validate(self):
        with tempfile.TemporaryDirectory() as directory:
            valid = Path(directory) / "valid.csv"
            with valid.open("w", encoding="utf-8", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=comparison.OUTPUT_FIELDS)
                writer.writeheader()
                writer.writerow(comparison_row())
            self.assertEqual(len(spreadsheet.read_comparisons([valid])), 1)

            invalid = Path(directory) / "invalid.csv"
            invalid.write_text("Chip\ngfx90a\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "missing columns"):
                spreadsheet.read_comparisons([invalid])

    def test_sheet_names_and_numbers(self):
        used = set()
        self.assertEqual(spreadsheet.sanitize_sheet_name("gfx:90a", used), "gfx_90a")
        self.assertEqual(spreadsheet.sanitize_sheet_name("gfx:90a", used), "gfx_90a_1")
        self.assertEqual(spreadsheet.sanitize_sheet_name("'", used), "unknown")
        long_name = spreadsheet.sanitize_sheet_name("x" * 40, used)
        self.assertEqual(len(long_name), 31)
        self.assertEqual(spreadsheet._number("1.5"), 1.5)
        self.assertIsNone(spreadsheet._number("bad"))
        self.assertIsNone(spreadsheet._number(None))
        self.assertIsNone(spreadsheet._number(str(math.inf)))

    def test_create_workbook_colors_and_layout(self):
        rows = [
            comparison_row(config="-g 1", diff=10),
            comparison_row(config="-g 2", diff=-10),
            comparison_row(config="-g 3", diff=0.1),
            comparison_row(config="-g 4", diff=float("nan")),
            comparison_row(chip="gfx1201", config="-g 1", diff=5),
            comparison_row(chip="gfx1201", config="-g 2", diff=5),
            comparison_row(chip="gfx1201", config="-g 3", diff=5),
            comparison_row(chip="gfx1201", config="-g 4", diff=5),
        ]
        workbook = spreadsheet.create_workbook(rows)
        self.assertEqual(workbook.sheetnames, ["gfx1201", "gfx90a"])
        sheet = workbook["gfx90a"]
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertTrue(sheet.auto_filter.ref)
        base_column = comparison.OUTPUT_FIELDS.index("Base TFlops") + 1
        candidate_column = comparison.OUTPUT_FIELDS.index("Candidate TFlops") + 1
        diff_column = comparison.OUTPUT_FIELDS.index("% Diff") + 1
        fills = {
            sheet.cell(row, diff_column).fill.fgColor.rgb for row in range(2, sheet.max_row + 1)
        }
        self.assertEqual(len(fills), 4)
        self.assertNotEqual(
            sheet.cell(2, base_column).fill.fgColor.rgb,
            sheet.cell(2, candidate_column).fill.fgColor.rgb)

        with self.assertRaisesRegex(ValueError, "No comparison"):
            spreadsheet.create_workbook([])
        with self.assertRaisesRegex(ValueError, "nonnegative"):
            spreadsheet.create_workbook(rows, -1)
        with self.assertRaisesRegex(ValueError, "Duplicate"):
            spreadsheet.create_workbook([rows[0], rows[0]])
        reordered = comparison_row(config="--g=1")
        with self.assertRaisesRegex(ValueError, "Duplicate"):
            spreadsheet.create_workbook([rows[0], reordered])
        mixed_revision = [comparison_row(), comparison_row(chip="gfx950")]
        mixed_revision[1]["Candidate SHA"] = "other"
        with self.assertRaisesRegex(ValueError, "revision pair"):
            spreadsheet.create_workbook(mixed_revision)
        with self.assertRaisesRegex(ValueError, "same configs"):
            spreadsheet.create_workbook(
                [comparison_row(), comparison_row(chip="gfx950", config="-g 2")])
        mismatched_flags = [comparison_row(), comparison_row(chip="gfx950")]
        mismatched_flags[1]["RocmlirGenFlags"] = "-current_seq_len=3"
        with self.assertRaisesRegex(ValueError, "runtime flags"):
            spreadsheet.create_workbook(mismatched_flags)
        bad_counts = comparison_row()
        bad_counts["Candidate Samples"] = "2"
        with self.assertRaisesRegex(ValueError, "sample counts"):
            spreadsheet.create_workbook([bad_counts])
        bad_counts["Candidate Samples"] = "bad"
        with self.assertRaisesRegex(ValueError, "positive integers"):
            spreadsheet.create_workbook([bad_counts])

    def test_save_cleanup_and_main(self):
        workbook = spreadsheet.create_workbook([comparison_row()])
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            output = root / "nested" / "result.xlsx"
            spreadsheet.save_workbook(workbook, output)
            self.assertEqual(load_workbook(output).sheetnames, ["gfx90a"])

            with mock.patch.object(workbook, "save", side_effect=OSError("save")):
                with self.assertRaisesRegex(OSError, "save"):
                    spreadsheet.save_workbook(workbook, root / "failure.xlsx")
            self.assertFalse(any(path.name.startswith(".failure") for path in root.iterdir()))

            with mock.patch.object(workbook, "save", side_effect=OSError("save")), \
                    mock.patch.object(os, "unlink", side_effect=FileNotFoundError):
                with self.assertRaisesRegex(OSError, "save"):
                    spreadsheet.save_workbook(workbook, root / "failure.xlsx")

            csv_path = root / "comparison.csv"
            with csv_path.open("w", encoding="utf-8", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=comparison.OUTPUT_FIELDS)
                writer.writeheader()
                writer.writerow(comparison_row())
            with mock.patch("builtins.print") as printed:
                self.assertEqual(
                    spreadsheet.main([
                        "--input",
                        str(csv_path),
                        "--output",
                        str(output),
                    ]), 0)
            printed.assert_called_once()


if __name__ == "__main__":
    unittest.main()
